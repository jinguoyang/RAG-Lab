"""生成RAG测试用的Excel和图片文件"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def create_progress_excel():
    """创建项目进度表"""
    wb = Workbook()

    # Sheet 1: 项目概览
    ws1 = wb.active
    ws1.title = "项目概览"

    # 标题样式
    title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题
    ws1.merge_cells('A1:F1')
    ws1['A1'] = '智云协同办公平台 V3.0 - 项目进度表'
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 40

    # 项目基本信息
    info_data = [
        ['项目名称', '智云协同办公平台 V3.0'],
        ['项目编号', 'ZY-2026-Q2-001'],
        ['项目负责人', '张明远'],
        ['起止日期', '2026-04-01 至 2026-06-30'],
        ['项目预算', '¥2,800,000'],
        ['当前进度', '68%'],
    ]

    for i, (key, value) in enumerate(info_data, start=3):
        ws1[f'A{i}'] = key
        ws1[f'A{i}'].font = Font(name='微软雅黑', size=11, bold=True)
        ws1[f'B{i}'] = value
        ws1[f'B{i}'].font = Font(name='微软雅黑', size=11)

    # Sheet 2: 迭代进度
    ws2 = wb.create_sheet("迭代进度")

    headers = ['迭代', '开始日期', '结束日期', '计划功能', '已完成', '进行中', '待开发', '完成率', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    sprints = [
        ['Sprint 1', '2026-04-01', '2026-04-14', 12, 12, 0, 0, '100%', '已完成'],
        ['Sprint 2', '2026-04-15', '2026-04-28', 15, 15, 0, 0, '100%', '已完成'],
        ['Sprint 3', '2026-04-29', '2026-05-12', 18, 16, 2, 0, '89%', '已完成'],
        ['Sprint 4', '2026-05-13', '2026-05-26', 20, 14, 4, 2, '70%', '进行中'],
        ['Sprint 5', '2026-05-27', '2026-06-09', 16, 0, 0, 16, '0%', '待开始'],
        ['Sprint 6', '2026-06-10', '2026-06-23', 10, 0, 0, 10, '0%', '待开始'],
        ['Sprint 7', '2026-06-24', '2026-06-30', 5, 0, 0, 5, '0%', '待开始'],
    ]

    for row_idx, sprint in enumerate(sprints, 2):
        for col_idx, value in enumerate(sprint, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            # 状态颜色
            if col_idx == 9:
                if value == '已完成':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value == '进行中':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Sheet 3: 任务明细
    ws3 = wb.create_sheet("任务明细")

    task_headers = ['任务ID', '所属迭代', '任务名称', '负责人', '优先级', '状态', '预估工时', '实际工时', '截止日期']
    for col, header in enumerate(task_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    tasks = [
        ['T-001', 'Sprint 1', '用户注册登录功能', '赵晓峰', 'P0', '已完成', 24, 22, '2026-04-07'],
        ['T-002', 'Sprint 1', '组织架构管理', '赵晓峰', 'P0', '已完成', 16, 18, '2026-04-10'],
        ['T-003', 'Sprint 1', '权限管理模块', '赵晓峰', 'P0', '已完成', 20, 24, '2026-04-14'],
        ['T-004', 'Sprint 2', '文档CRUD接口', '赵晓峰', 'P0', '已完成', 32, 30, '2026-04-21'],
        ['T-005', 'Sprint 2', '文档版本管理', '赵晓峰', 'P0', '已完成', 24, 28, '2026-04-25'],
        ['T-006', 'Sprint 2', '文件上传功能', '赵晓峰', 'P0', '已完成', 16, 16, '2026-04-28'],
        ['T-007', 'Sprint 3', '在线编辑器集成', '陈思远', 'P0', '已完成', 40, 44, '2026-05-05'],
        ['T-008', 'Sprint 3', '协同编辑基础功能', '陈思远', 'P0', '已完成', 32, 36, '2026-05-10'],
        ['T-009', 'Sprint 3', '评论批注功能', '陈思远', 'P1', '已完成', 16, 14, '2026-05-12'],
        ['T-010', 'Sprint 4', '协同编辑冲突修复', '陈思远', 'P0', '进行中', 24, 16, '2026-05-25'],
        ['T-011', 'Sprint 4', 'AI问答基础框架', '李建国', 'P0', '已完成', 40, 38, '2026-05-18'],
        ['T-012', 'Sprint 4', 'AI问答知识库对接', '李建国', 'P0', '进行中', 32, 20, '2026-05-28'],
        ['T-013', 'Sprint 4', '消息推送优化', '赵晓峰', 'P1', '进行中', 16, 10, '2026-05-25'],
        ['T-014', 'Sprint 4', '任务看板功能', '陈思远', 'P1', '进行中', 24, 8, '2026-05-30'],
        ['T-015', 'Sprint 4', '性能测试', '刘芳华', 'P0', '待开始', 24, 0, '2026-06-01'],
        ['T-016', 'Sprint 4', '文档模板功能', '赵晓峰', 'P2', '待开始', 16, 0, '2026-06-03'],
    ]

    for row_idx, task in enumerate(tasks, 2):
        for col_idx, value in enumerate(task, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            # 状态颜色
            if col_idx == 6:
                if value == '已完成':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value == '进行中':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # 调整列宽
    for ws in [ws1, ws2, ws3]:
        for col_cells in ws.columns:
            max_length = 0
            # 获取列字母 - 处理合并单元格
            column = None
            for cell in col_cells:
                try:
                    column = cell.column_letter
                    break
                except:
                    continue
            if column is None:
                continue
            for cell in col_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(os.path.join(OUTPUT_DIR, '05-项目进度表.xlsx'))
    print("[OK] 项目进度表已生成")


def create_employee_excel():
    """创建人员花名册"""
    wb = Workbook()
    ws = wb.active
    ws.title = "人员花名册"

    # 样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells('A1:I1')
    ws['A1'] = '智云协同办公平台 V3.0 - 项目组人员花名册'
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # 表头
    headers = ['工号', '姓名', '部门', '职位', '项目角色', '联系方式', '入职日期', '技术栈', '负责模块']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # 人员数据
    employees = [
        ['ZY001', '张明远', '项目管理部', '高级项目经理', '项目经理', 'zhangmy@zhiyun.com', '2020-03-15', 'PMP, Agile', '项目整体管理'],
        ['ZY002', '李建国', '研发部', '技术总监', '技术负责人', 'lijg@zhiyun.com', '2018-07-01', 'Java, Python, Go, K8s', '架构设计, AI服务'],
        ['ZY003', '王雪婷', '产品部', '高级产品经理', '产品经理', 'wangxt@zhiyun.com', '2021-01-10', 'Axure, Figma', '需求管理, 产品设计'],
        ['ZY004', '陈思远', '前端组', '前端工程师', '前端负责人', 'chensy@zhiyun.com', '2022-03-20', 'React, TypeScript, Vue', '文档协作, 任务管理'],
        ['ZY005', '赵晓峰', '后端组', '后端工程师', '后端负责人', 'zhaoxf@zhiyun.com', '2021-06-15', 'Java, Spring Boot, MyBatis', '用户服务, 文档服务'],
        ['ZY006', '刘芳华', '测试部', '测试工程师', '测试负责人', 'liufh@zhiyun.com', '2022-09-01', 'Selenium, JMeter, Pytest', '功能测试, 性能测试'],
        ['ZY007', '周大伟', '运维部', '运维工程师', '运维工程师', 'zhoudw@zhiyun.com', '2023-02-20', 'Docker, K8s, Jenkins', '部署运维'],
        ['ZY008', '孙小明', '前端组', '前端工程师', '前端开发', 'sunxm@zhiyun.com', '2023-07-10', 'React, CSS, Webpack', '用户管理前端'],
        ['ZY009', '钱丽华', '后端组', '后端工程师', '后端开发', 'qianlh@zhiyun.com', '2023-09-15', 'Java, PostgreSQL, Redis', '消息服务'],
        ['ZY010', '吴志强', '研发部', 'AI工程师', 'AI开发', 'wuzq@zhiyun.com', '2024-01-08', 'Python, PyTorch, LLM', 'AI助手开发'],
        ['ZY011', '郑雅琴', '测试部', '测试工程师', '测试开发', 'zhengyq@zhiyun.com', '2024-03-20', 'Python, Selenium, CI/CD', '自动化测试'],
        ['ZY012', '林海涛', '前端组', '初级前端工程师', '前端开发', 'linht@zhiyun.com', '2025-06-01', 'React, JavaScript', '即时通讯前端'],
    ]

    for row_idx, emp in enumerate(employees, 4):
        for col_idx, value in enumerate(emp, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # 调整列宽
    col_widths = [10, 12, 15, 18, 15, 25, 15, 30, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(os.path.join(OUTPUT_DIR, '06-人员花名册.xlsx'))
    print("[OK] 人员花名册已生成")


def create_architecture_diagram():
    """创建系统架构图"""
    fig, ax = plt.subplots(1, 1, figsize=(14, 10))

    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False

    # 标题
    ax.set_title('智云协同办公平台 V3.0 - 系统架构图', fontsize=18, fontweight='bold', pad=20)

    # 定义层级
    layers = [
        {'y': 8.5, 'height': 1.2, 'color': '#E8F4FD', 'border': '#2196F3', 'label': '接入层',
         'items': [('Nginx\n负载均衡', 2), ('Kong\nAPI网关', 5), ('CDN\n内容分发', 8), ('WAF\n防火墙', 11)]},
        {'y': 6.5, 'height': 1.5, 'color': '#E8F5E9', 'border': '#4CAF50', 'label': '应用层',
         'items': [('用户服务\n8001', 1), ('文档服务\n8002', 3.5), ('消息服务\n8003', 6), ('AI服务\n8004', 8.5), ('任务服务\n8005', 11)]},
        {'y': 4.5, 'height': 1.5, 'color': '#FFF3E0', 'border': '#FF9800', 'label': '数据层',
         'items': [('PostgreSQL\n主数据库', 1.5), ('Redis\n缓存', 4.5), ('MinIO\n文件存储', 7.5), ('Elasticsearch\n搜索引擎', 10.5)]},
        {'y': 2.5, 'height': 1.5, 'color': '#F3E5F5', 'border': '#9C27B0', 'label': '基础设施层',
         'items': [('Docker\n容器', 1.5), ('Kubernetes\n编排', 4.5), ('Prometheus\n监控', 7.5), ('Grafana\n可视化', 10.5)]},
    ]

    for layer in layers:
        # 背景矩形
        rect = plt.Rectangle((0.3, layer['y']), 12.4, layer['height'],
                            facecolor=layer['color'], edgecolor=layer['border'],
                            linewidth=2, alpha=0.8)
        ax.add_patch(rect)

        # 层级标签
        ax.text(0.5, layer['y'] + layer['height']/2, layer['label'],
                fontsize=12, fontweight='bold', va='center', ha='left',
                color=layer['border'])

        # 组件
        for (label, x) in layer['items']:
            comp_rect = plt.Rectangle((x-0.6, layer['y']+0.15), 1.8, layer['height']*0.7,
                                     facecolor='white', edgecolor=layer['border'],
                                     linewidth=1.5, alpha=0.9)
            ax.add_patch(comp_rect)
            ax.text(x+0.3, layer['y']+layer['height']/2, label,
                    fontsize=8, va='center', ha='center')

    # 连接箭头
    arrow_props = dict(arrowstyle='->', color='#666666', lw=1.5)
    for y_start, y_end in [(8.5, 8.0), (6.5, 6.0), (4.5, 4.0)]:
        ax.annotate('', xy=(6.5, y_end), xytext=(6.5, y_start),
                    arrowprops=arrow_props)

    # 前端应用标注
    ax.text(6.5, 10.2, 'Web前端 (React) | Mobile (React Native) | 桌面端 (Electron)',
            fontsize=10, ha='center', va='center',
            bbox=dict(boxstyle='round,pad=0.5', facecolor='#BBDEFB', edgecolor='#1976D2'))

    ax.set_xlim(0, 13)
    ax.set_ylim(1.5, 11)
    ax.axis('off')

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '08-系统架构图.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("[OK] 系统架构图已生成")


def create_prototype_diagram():
    """创建产品原型图"""
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))

    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False

    fig.suptitle('智云协同办公平台 V3.0 - 产品原型图', fontsize=18, fontweight='bold')

    # 原型1: 文档列表页
    ax1 = axes[0, 0]
    ax1.set_title('文档列表页', fontsize=14, fontweight='bold')

    # 顶部导航栏
    nav_rect = plt.Rectangle((0, 0.88), 1, 0.12, facecolor='#1976D2', transform=ax1.transAxes)
    ax1.add_patch(nav_rect)
    ax1.text(0.02, 0.94, '智云协同办公', color='white', fontsize=10, transform=ax1.transAxes)

    # 搜索框
    search_rect = plt.Rectangle((0.25, 0.89), 0.5, 0.06, facecolor='white', edgecolor='#ccc',
                                transform=ax1.transAxes)
    ax1.add_patch(search_rect)
    ax1.text(0.35, 0.92, '搜索文档...', color='#999', fontsize=9, transform=ax1.transAxes)

    # 侧边栏
    sidebar_rect = plt.Rectangle((0, 0), 0.2, 0.88, facecolor='#F5F5F5', edgecolor='#E0E0E0',
                                transform=ax1.transAxes)
    ax1.add_patch(sidebar_rect)
    menu_items = ['全部文档', '我创建的', '共享给我', '最近编辑', '已归档']
    for i, item in enumerate(menu_items):
        y = 0.82 - i * 0.08
        color = '#1976D2' if i == 0 else '#333'
        ax1.text(0.03, y, f'📁 {item}', fontsize=9, color=color, transform=ax1.transAxes)

    # 文档列表
    docs = [
        ('产品需求规格说明书', '王雪婷', '2026-05-18', '已发布'),
        ('技术架构设计文档', '李建国', '2026-05-20', '已发布'),
        ('第五次迭代会议纪要', '王雪婷', '2026-05-20', '草稿'),
        ('测试报告 V1.0', '刘芳华', '2026-05-15', '审核中'),
        ('接口规范文档', '赵晓峰', '2026-05-19', '已发布'),
    ]

    # 表头
    ax1.text(0.22, 0.83, '文档名称', fontsize=9, fontweight='bold', transform=ax1.transAxes)
    ax1.text(0.55, 0.83, '作者', fontsize=9, fontweight='bold', transform=ax1.transAxes)
    ax1.text(0.7, 0.83, '更新时间', fontsize=9, fontweight='bold', transform=ax1.transAxes)
    ax1.text(0.88, 0.83, '状态', fontsize=9, fontweight='bold', transform=ax1.transAxes)
    ax1.plot([0.2, 1], [0.82, 0.82], color='#E0E0E0', transform=ax1.transAxes, clip_on=False)

    for i, (name, author, date, status) in enumerate(docs):
        y = 0.77 - i * 0.08
        ax1.text(0.22, y, name, fontsize=8, transform=ax1.transAxes)
        ax1.text(0.55, y, author, fontsize=8, transform=ax1.transAxes)
        ax1.text(0.7, y, date, fontsize=8, transform=ax1.transAxes)
        status_color = '#4CAF50' if status == '已发布' else '#FF9800' if status == '审核中' else '#9E9E9E'
        ax1.text(0.88, y, status, fontsize=8, color=status_color, transform=ax1.transAxes)
        ax1.plot([0.2, 1], [y-0.03, y-0.03], color='#F0F0F0', transform=ax1.transAxes, clip_on=False)

    ax1.axis('off')

    # 原型2: 文档编辑页
    ax2 = axes[0, 1]
    ax2.set_title('文档编辑页', fontsize=14, fontweight='bold')

    # 工具栏
    toolbar_rect = plt.Rectangle((0, 0.9), 1, 0.1, facecolor='#FAFAFA', edgecolor='#E0E0E0',
                                transform=ax2.transAxes)
    ax2.add_patch(toolbar_rect)
    tools = ['B', 'I', 'U', 'H1', 'H2', '🔗', '📷', '📊']
    for i, tool in enumerate(tools):
        ax2.text(0.05 + i * 0.08, 0.95, tool, fontsize=10, fontweight='bold',
                transform=ax2.transAxes,
                bbox=dict(boxstyle='round', facecolor='white', edgecolor='#DDD'))

    # 协作者头像
    ax2.text(0.75, 0.95, '👤👤👤 在线协作中', fontsize=8, color='#4CAF50', transform=ax2.transAxes)

    # 编辑区域
    edit_rect = plt.Rectangle((0.02, 0.05), 0.96, 0.83, facecolor='white', edgecolor='#E0E0E0',
                             transform=ax2.transAxes)
    ax2.add_patch(edit_rect)

    content_lines = [
        '# 技术架构设计文档',
        '',
        '## 1. 系统架构概述',
        '',
        '本系统采用微服务架构设计，主要包含...',
        '',
        '### 1.1 核心服务',
        '',
        '| 服务名称 | 端口 | 负责人 |',
        '|----------|------|--------|',
        '| user-service | 8001 | 赵晓峰 |',
        '| document-service | 8002 | 赵晓峰 |',
        '',
        '## 2. 数据库设计',
        '...',
    ]

    for i, line in enumerate(content_lines):
        y = 0.85 - i * 0.05
        fontsize = 12 if line.startswith('# ') else 10 if line.startswith('## ') else 9
        fontweight = 'bold' if line.startswith('#') else 'normal'
        ax2.text(0.05, y, line, fontsize=fontsize, fontweight=fontweight, transform=ax2.transAxes)

    ax2.axis('off')

    # 原型3: 即时通讯页
    ax3 = axes[1, 0]
    ax3.set_title('即时通讯页', fontsize=14, fontweight='bold')

    # 左侧会话列表
    conv_rect = plt.Rectangle((0, 0), 0.3, 1, facecolor='#F5F5F5', edgecolor='#E0E0E0',
                             transform=ax3.transAxes)
    ax3.add_patch(conv_rect)

    ax3.text(0.02, 0.96, '消息', fontsize=12, fontweight='bold', transform=ax3.transAxes)

    conversations = [
        ('项目组群聊', '李建国: AI方案已更新...', '14:30'),
        ('张明远', '会议纪要确认一下', '13:45'),
        ('技术讨论组', '赵晓峰: 分片上传已完成', '12:20'),
        ('陈思远', '协同编辑的bug已修复', '11:05'),
    ]

    for i, (name, last_msg, time) in enumerate(conversations):
        y = 0.88 - i * 0.15
        ax3.text(0.02, y, name, fontsize=9, fontweight='bold', transform=ax3.transAxes)
        ax3.text(0.02, y-0.04, last_msg, fontsize=7, color='#666', transform=ax3.transAxes)
        ax3.text(0.25, y, time, fontsize=7, color='#999', transform=ax3.transAxes)

    # 右侧聊天区域
    chat_rect = plt.Rectangle((0.3, 0.08), 0.7, 0.92, facecolor='white', transform=ax3.transAxes)
    ax3.add_patch(chat_rect)

    # 聊天消息
    messages = [
        ('张明远', '今天的会议纪要大家确认一下', '14:25', False),
        ('李建国', '好的，AI部分我补充一下CRDT方案的进度', '14:28', False),
        ('李建国', '已经更新到文档里了，请查看', '14:30', True),
        ('我', '收到，我看一下', '14:32', None),
    ]

    for i, (sender, msg, time, is_right) in enumerate(messages):
        y = 0.85 - i * 0.12
        if is_right is None:  # 自己的消息
            ax3.text(0.95, y, msg, fontsize=8, ha='right', transform=ax3.transAxes,
                    bbox=dict(boxstyle='round', facecolor='#DCF8C6', edgecolor='#4CAF50'))
        elif is_right:
            ax3.text(0.35, y, f'{sender}: {msg}', fontsize=8, transform=ax3.transAxes,
                    bbox=dict(boxstyle='round', facecolor='#E3F2FD', edgecolor='#2196F3'))
        else:
            ax3.text(0.35, y, f'{sender}: {msg}', fontsize=8, transform=ax3.transAxes,
                    bbox=dict(boxstyle='round', facecolor='#F5F5F5', edgecolor='#E0E0E0'))

    # 输入框
    input_rect = plt.Rectangle((0.3, 0), 0.6, 0.07, facecolor='white', edgecolor='#E0E0E0',
                              transform=ax3.transAxes)
    ax3.add_patch(input_rect)
    ax3.text(0.32, 0.035, '输入消息...', color='#999', fontsize=9, transform=ax3.transAxes)

    ax3.axis('off')

    # 原型4: 任务看板页
    ax4 = axes[1, 1]
    ax4.set_title('任务看板页', fontsize=14, fontweight='bold')

    # 看板列
    columns = [
        ('待办', '#FF9800', ['编写单元测试', '设计文档模板', '性能优化方案']),
        ('进行中', '#2196F3', ['协同编辑修复', 'AI问答优化', '消息推送优化']),
        ('测试中', '#9C27B0', ['用户导入功能', '文件上传功能']),
        ('已完成', '#4CAF50', ['用户登录注册', '文档CRUD', '版本管理', '权限管理']),
    ]

    for i, (title, color, tasks) in enumerate(columns):
        x = 0.02 + i * 0.245

        # 列头
        header_rect = plt.Rectangle((x, 0.88), 0.23, 0.1, facecolor=color, transform=ax4.transAxes)
        ax4.add_patch(header_rect)
        ax4.text(x + 0.115, 0.93, f'{title} ({len(tasks)})', fontsize=9, fontweight='bold',
                color='white', ha='center', transform=ax4.transAxes)

        # 任务卡片
        for j, task in enumerate(tasks):
            y = 0.78 - j * 0.15
            card_rect = plt.Rectangle((x + 0.01, y), 0.21, 0.12, facecolor='white',
                                     edgecolor='#E0E0E0', transform=ax4.transAxes)
            ax4.add_patch(card_rect)
            ax4.text(x + 0.03, y + 0.08, task, fontsize=8, transform=ax4.transAxes)
            # 优先级标签
            priority = 'P0' if j == 0 else 'P1' if j == 1 else 'P2'
            p_color = '#F44336' if priority == 'P0' else '#FF9800' if priority == 'P1' else '#4CAF50'
            ax4.text(x + 0.18, y + 0.02, priority, fontsize=7, color=p_color,
                    transform=ax4.transAxes)

    ax4.axis('off')

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '09-产品原型图.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("[OK] 产品原型图已生成")


if __name__ == '__main__':
    create_progress_excel()
    create_employee_excel()
    create_architecture_diagram()
    create_prototype_diagram()
    print("\n[OK] 所有测试文件生成完成！")
